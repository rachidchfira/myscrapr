from __future__ import annotations

import csv
import io
import json
import os
import re
import uuid
from collections.abc import Callable, Sequence
from contextlib import suppress
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from mapslead.config import Settings
from mapslead.errors import ExportError
from mapslead.models import ExportPaths, RunSnapshot
from mapslead.normalize import normalize_text
from mapslead.ports import RepositoryPort

_CSV_FIELDNAMES = (
    "place_id",
    "name",
    "category",
    "address",
    "phone",
    "website",
    "rating",
    "review_count",
    "google_maps_url",
    "emails",
    "facebook_url",
    "instagram_url",
    "linkedin_url",
    "x_url",
    "youtube_url",
    "business_type",
    "location_query",
    "first_seen_at",
    "last_seen_at",
    "enrichment_status",
    "enrichment_error",
    "run_id",
)

_WORKBOOK_TIMESTAMP = datetime(2000, 1, 1, 0, 0, 0, tzinfo=UTC)
_ZIP_ENTRY_TIMESTAMP = _WORKBOOK_TIMESTAMP.timetuple()[:6]
_WORKBOOK_TIMESTAMP_XML = _WORKBOOK_TIMESTAMP.strftime("%Y-%m-%dT%H:%M:%SZ")
_MAX_COLUMN_WIDTH = 60
_CORE_XML_MODIFIED_PATTERN = re.compile(
    rb"(<dcterms:modified xsi:type=\"dcterms:W3CDTF\">)([^<]+)(</dcterms:modified>)"
)
type _CellValue = str | int | float | None


@dataclass(frozen=True, slots=True)
class _PreparedRow:
    business_id: int
    run_id: str
    name: str
    business_type: str
    location_query: str
    first_seen_at: str
    last_seen_at: str
    place_id: str | None
    category: str | None
    address: str | None
    phone: str | None
    website: str | None
    rating: float | None
    review_count: int | None
    google_maps_url: str | None
    emails: tuple[str, ...]
    facebook_url: str | None
    instagram_url: str | None
    linkedin_url: str | None
    x_url: str | None
    youtube_url: str | None
    enrichment_status: str
    enrichment_error: str | None


@dataclass(frozen=True, slots=True)
class _ReplaceTarget:
    temp_path: Path
    destination_path: Path
    backup_path: Path
    destination_existed: bool


class Exporter:
    def __init__(self, repository: RepositoryPort, settings: Settings) -> None:
        self._repository = repository
        self._settings = settings

    def export_run(self, run_id: str) -> ExportPaths:
        run_dir = _validated_run_dir(self._settings.export_dir, run_id)
        snapshots = tuple(
            sorted(
                self._repository.snapshots_for_run(run_id),
                key=lambda snapshot: (
                    normalize_text(snapshot.name),
                    normalize_text(snapshot.address),
                    snapshot.business_id,
                ),
            )
        )
        rows = tuple(_prepare_row(snapshot) for snapshot in snapshots)

        run_dir.mkdir(parents=True, exist_ok=True)

        csv_path = run_dir / "results.csv"
        json_path = run_dir / "results.json"
        xlsx_path = run_dir / "results.xlsx"
        csv_temp_path = run_dir / "results.csv.tmp"
        json_temp_path = run_dir / "results.json.tmp"
        xlsx_temp_path = run_dir / "results.xlsx.tmp"

        try:
            csv_document = _build_csv_document(rows)
            _write_atomic_text(csv_temp_path, csv_document)

            json_document = _build_json_document(rows)
            _write_atomic_text(json_temp_path, json_document)

            xlsx_document = _build_xlsx_document(rows, _CSV_FIELDNAMES, _csv_row)
            _write_atomic_bytes(xlsx_temp_path, xlsx_document)

            _replace_export_targets(
                (
                    _replace_target(csv_temp_path, csv_path),
                    _replace_target(json_temp_path, json_path),
                    _replace_target(xlsx_temp_path, xlsx_path),
                )
            )
        except Exception as error:
            _cleanup_path(csv_temp_path)
            _cleanup_path(json_temp_path)
            _cleanup_path(xlsx_temp_path)
            raise ExportError(f"failed to export run {run_id}") from error

        return ExportPaths(csv_path=csv_path, json_path=json_path, xlsx_path=xlsx_path)


def _prepare_row(snapshot: RunSnapshot) -> _PreparedRow:
    return _PreparedRow(
        business_id=snapshot.business_id,
        run_id=snapshot.run_id,
        name=snapshot.name,
        business_type=snapshot.business_type,
        location_query=snapshot.location_query,
        first_seen_at=snapshot.first_seen_at.isoformat(),
        last_seen_at=snapshot.last_seen_at.isoformat(),
        place_id=snapshot.place_id,
        category=snapshot.category,
        address=snapshot.address,
        phone=snapshot.phone,
        website=snapshot.website,
        rating=snapshot.rating,
        review_count=snapshot.review_count,
        google_maps_url=snapshot.google_maps_url,
        emails=tuple(sorted(snapshot.emails)),
        facebook_url=snapshot.facebook_url,
        instagram_url=snapshot.instagram_url,
        linkedin_url=snapshot.linkedin_url,
        x_url=snapshot.x_url,
        youtube_url=snapshot.youtube_url,
        enrichment_status=snapshot.enrichment_status.value,
        enrichment_error=snapshot.enrichment_error,
    )


def _build_csv_document(rows: tuple[_PreparedRow, ...]) -> str:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=_CSV_FIELDNAMES, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(_csv_row(row))
    return buffer.getvalue()


def _csv_row(row: _PreparedRow) -> dict[str, str | int | float | None]:
    return {
        "place_id": row.place_id,
        "name": row.name,
        "category": row.category,
        "address": row.address,
        "phone": row.phone,
        "website": row.website,
        "rating": row.rating,
        "review_count": row.review_count,
        "google_maps_url": row.google_maps_url,
        "emails": ";".join(row.emails),
        "facebook_url": row.facebook_url,
        "instagram_url": row.instagram_url,
        "linkedin_url": row.linkedin_url,
        "x_url": row.x_url,
        "youtube_url": row.youtube_url,
        "business_type": row.business_type,
        "location_query": row.location_query,
        "first_seen_at": row.first_seen_at,
        "last_seen_at": row.last_seen_at,
        "enrichment_status": row.enrichment_status,
        "enrichment_error": row.enrichment_error,
        "run_id": row.run_id,
    }


def _build_json_document(rows: tuple[_PreparedRow, ...]) -> str:
    return json.dumps(
        [_json_row(row) for row in rows],
        ensure_ascii=False,
        indent=2,
    ) + "\n"


def _json_row(row: _PreparedRow) -> dict[str, str | int | float | list[str] | None]:
    return {
        "place_id": row.place_id,
        "name": row.name,
        "category": row.category,
        "address": row.address,
        "phone": row.phone,
        "website": row.website,
        "rating": row.rating,
        "review_count": row.review_count,
        "google_maps_url": row.google_maps_url,
        "emails": list(row.emails),
        "facebook_url": row.facebook_url,
        "instagram_url": row.instagram_url,
        "linkedin_url": row.linkedin_url,
        "x_url": row.x_url,
        "youtube_url": row.youtube_url,
        "business_type": row.business_type,
        "location_query": row.location_query,
        "first_seen_at": row.first_seen_at,
        "last_seen_at": row.last_seen_at,
        "enrichment_status": row.enrichment_status,
        "enrichment_error": row.enrichment_error,
        "run_id": row.run_id,
    }


def _build_xlsx_document[
    RowT
](
    rows: Sequence[RowT],
    fieldnames: Sequence[str],
    row_builder: Callable[[RowT], dict[str, _CellValue]],
) -> bytes:
    workbook = Workbook()
    workbook.properties.created = _WORKBOOK_TIMESTAMP
    workbook.properties.modified = _WORKBOOK_TIMESTAMP
    workbook.properties.lastModifiedBy = "mapslead"
    workbook.properties.creator = "mapslead"

    worksheet = workbook.active
    worksheet.title = "Results"
    worksheet.freeze_panes = "A2"

    column_widths = [len(fieldname) for fieldname in fieldnames]
    for column_index, fieldname in enumerate(fieldnames, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=fieldname)
        cell.data_type = "s"

    for row_index, row in enumerate(rows, start=2):
        values = row_builder(row)
        for column_index, fieldname in enumerate(fieldnames, start=1):
            value = values[fieldname]
            cell = worksheet.cell(row=row_index, column=column_index)
            _assign_cell_value(cell, value)
            display_value = "" if value is None else str(value)
            column_widths[column_index - 1] = max(column_widths[column_index - 1], len(display_value))

    last_column = get_column_letter(len(fieldnames))
    worksheet.auto_filter.ref = f"A1:{last_column}{max(len(rows) + 1, 1)}"
    for column_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(width + 2, _MAX_COLUMN_WIDTH)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return _canonicalize_zip_archive(buffer.getvalue())


def _canonicalize_zip_archive(document: bytes) -> bytes:
    source_buffer = io.BytesIO(document)
    output_buffer = io.BytesIO()
    with ZipFile(source_buffer) as source_archive:
        entries = {
            entry.filename: _canonicalize_zip_payload(entry.filename, source_archive.read(entry.filename))
            for entry in sorted(source_archive.infolist(), key=lambda entry: entry.filename)
        }

    with ZipFile(output_buffer, "w", compression=ZIP_DEFLATED, compresslevel=9) as output_archive:
        output_archive.comment = b""
        for filename, payload in entries.items():
            zip_info = ZipInfo(filename=filename, date_time=_ZIP_ENTRY_TIMESTAMP)
            zip_info.compress_type = ZIP_DEFLATED
            zip_info.create_system = 0
            zip_info.create_version = 20
            zip_info.extract_version = 20
            zip_info.flag_bits = 0
            zip_info.volume = 0
            zip_info.internal_attr = 0
            zip_info.external_attr = 0
            zip_info.extra = b""
            zip_info.comment = b""
            output_archive.writestr(zip_info, payload, compress_type=ZIP_DEFLATED, compresslevel=9)

    return output_buffer.getvalue()


def _canonicalize_zip_payload(filename: str, payload: bytes) -> bytes:
    if filename != "docProps/core.xml":
        return payload

    return _CORE_XML_MODIFIED_PATTERN.sub(
        rb"\g<1>" + _WORKBOOK_TIMESTAMP_XML.encode("ascii") + rb"\g<3>",
        payload,
        count=1,
    )


def _assign_cell_value(cell: Cell, value: _CellValue) -> None:
    if value is None:
        cell.value = None
        return
    if isinstance(value, bool):
        cell.value = str(value)
        cell.data_type = "s"
        return
    if isinstance(value, int):
        cell.value = value
        return
    if isinstance(value, float):
        cell.value = value
        return
    cell.value = value
    cell.data_type = "s"


def _write_atomic_text(path: Path, document: str) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        handle.write(document)
        handle.flush()
        os.fsync(handle.fileno())


def _write_atomic_bytes(path: Path, document: bytes) -> None:
    with path.open("wb") as handle:
        handle.write(document)
        handle.flush()
        os.fsync(handle.fileno())


def _replace_target(temp_path: Path, destination_path: Path) -> _ReplaceTarget:
    suffix = uuid.uuid4().hex
    return _ReplaceTarget(
        temp_path=temp_path,
        destination_path=destination_path,
        backup_path=destination_path.with_name(f"{destination_path.name}.{suffix}.bak"),
        destination_existed=destination_path.exists(),
    )


def _replace_export_targets(targets: tuple[_ReplaceTarget, ...]) -> None:
    prepared: list[_ReplaceTarget] = []
    replaced: list[_ReplaceTarget] = []
    if not targets:
        raise ValueError("export target group cannot be empty")
    directory = targets[0].destination_path.parent
    try:
        for target in targets:
            if target.destination_path.parent != directory:
                raise ValueError("export targets must share a directory")
            if target.destination_existed:
                os.replace(target.destination_path, target.backup_path)
            prepared.append(target)

        for target in targets:
            os.replace(target.temp_path, target.destination_path)
            replaced.append(target)

        _fsync_directory(directory)
    except Exception:
        for target in reversed(replaced):
            if target.destination_existed:
                with suppress(FileNotFoundError):
                    os.replace(target.backup_path, target.destination_path)
            else:
                _cleanup_path(target.destination_path)

        for target in reversed(prepared):
            if target in replaced:
                continue
            if target.destination_existed:
                with suppress(FileNotFoundError):
                    os.replace(target.backup_path, target.destination_path)

        _fsync_directory(directory)
        raise
    finally:
        for target in targets:
            _cleanup_path(target.temp_path)
            _cleanup_path(target.backup_path)


def _cleanup_path(path: Path) -> None:
    with suppress(FileNotFoundError):
        path.unlink()


def _fsync_directory(path: Path) -> None:
    try:
        descriptor = os.open(path, os.O_RDONLY)
    except OSError:
        return

    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _validated_run_dir(export_root: Path, run_id: str) -> Path:
    if not run_id or run_id in {".", ".."}:
        raise ExportError(f"unsafe run_id: {run_id!r}")

    run_path = Path(run_id)
    if run_path.is_absolute() or "/" in run_id or "\\" in run_id or len(run_path.parts) != 1:
        raise ExportError(f"unsafe run_id: {run_id!r}")

    resolved_export_root = export_root.resolve(strict=False)
    resolved_run_dir = (resolved_export_root / run_id).resolve(strict=False)
    if resolved_run_dir.parent != resolved_export_root:
        raise ExportError(f"unsafe run_id: {run_id!r}")

    return resolved_run_dir
